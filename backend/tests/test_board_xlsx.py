"""보드 XLSX 내보내기 — plan.md §0.5.7.

**중심은 round-trip 계약이다.** 내보낸 파일을 `db/migrate.py` 의 `parse_workbook`
이 그대로 다시 읽어야 한다 — 내보내기가 곧 임포트 양식이기 때문이다. 표기를 하나만
흩뜨려도(`Phase 0 Pre` 처럼 점을 빠뜨리거나, 문서 구분자를 바꾸거나) 재파싱이
깨지므로, 그 테스트가 이 모듈의 fail-closed 게이트다.

나머지는 그 계약을 떠받치는 것들이다: 원본 양식과 같은 자리에 헤더가 있는가,
미배정 행이 빈 셀로 나가는가, 빈 보드가 터지지 않는가.
"""

from __future__ import annotations

import importlib.util
import re
import sys
from datetime import date
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import pytest

from app.models import Item, ItemDocument, ItemOwner, ItemStatus
from app.schemas.item import DocumentRef, ItemOut, OwnerRef
from app.services import board_xlsx_service
from app.services.board_xlsx_service import (
    BOARD_FIRST_DATA_ROW,
    BOARD_HEADER_ROW,
    BOARD_HEADERS,
    BOARD_SHEET,
    DOC_HEADER_ROW,
    DOC_HEADERS,
    DOC_SHEET,
    EXPORT_STATUS,
    BoardExport,
    build_board_xlsx,
    documents_cell,
    milestone_cell,
    owners_cell,
    phase_cell,
)

API = "/api/v1"
REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = REPO_ROOT / "docs" / "Work Package.xlsx"


def load_migrate():
    """`db/` 는 패키지가 아니므로 경로로 직접 로드한다."""
    spec = importlib.util.spec_from_file_location(
        "wp_migrate_xlsx", REPO_ROOT / "db" / "migrate.py"
    )
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


migrate = load_migrate()


def open_xlsx(payload: bytes):
    from openpyxl import load_workbook

    return load_workbook(BytesIO(payload))


# =============================================================================
# 셀 표기 — 순수 함수
# =============================================================================
def row(
    no: int,
    *,
    phase_id: int | None = None,
    phase_no: int | None = None,
    phase_name: str = "",
    milestone_id: int | None = None,
    milestone_no: int | None = None,
    milestone_name: str = "",
    status: ItemStatus = ItemStatus.NOT_STARTED,
    title: str | None = None,
    deliverable: str | None = None,
    documents: list[tuple[str, str]] | None = None,
    owners: list[str] | None = None,
    completion_date: date | None = None,
) -> ItemOut:
    return ItemOut(
        id=no, sort_order=no, row_no=no,
        phase_id=phase_id, phase_no=phase_no, phase_name=phase_name,
        milestone_id=milestone_id, milestone_no=milestone_no, milestone_name=milestone_name,
        is_phase_block_start=False, is_phase_block_end=False,
        is_milestone_block_start=False, is_milestone_block_end=False,
        can_create_phase=False, can_create_milestone=False,
        status=status, title=title, deliverable=deliverable, completion_date=completion_date,
        documents=[
            DocumentRef(id=i, no=int(c), name=n)
            for i, (c, n) in enumerate(documents or [], start=1)
        ],
        owners=[OwnerRef(id=i, name=n) for i, n in enumerate(owners or [], start=1)],
    )


def test_phase_and_milestone_use_the_original_notation():
    item = row(1, phase_id=1, phase_no=0, phase_name="Pre-Infrastructure Setup",
               milestone_id=10, milestone_no=1, milestone_name="DSEP 환경 Gap 및 자원 구성")

    assert phase_cell(item) == "Phase 0. Pre-Infrastructure Setup"
    assert milestone_cell(item) == "0.1 DSEP 환경 Gap 및 자원 구성"
    # 파서가 그대로 되읽는다 — 이 두 정규식이 표기의 근거다.
    assert migrate.PHASE_RE.match(phase_cell(item)).groups() == (
        "0", "Pre-Infrastructure Setup"
    )
    assert migrate.MILESTONE_RE.match(milestone_cell(item)).groups() == (
        "0", "1", "DSEP 환경 Gap 및 자원 구성"
    )


def test_an_unassigned_row_exports_empty_phase_and_milestone_cells():
    """§0.5.7 — 회색 행은 빈 셀이다 (원본 양식에는 없던 상태)."""
    assert phase_cell(row(1)) == ""
    assert milestone_cell(row(1)) == ""
    # Phase 만 있고 Milestone 이 없는 행도 마찬가지.
    partial = row(2, phase_id=1, phase_no=0, phase_name="P")
    assert phase_cell(partial) == "Phase 0. P"
    assert milestone_cell(partial) == ""


def test_documents_are_written_as_numbers():
    """§0.5.10 ③ — 표기가 **숫자**로 통일됐다. 원문자는 폐기.

    이름을 싣지 않으므로 `DSEP Readiness & I/O Spec` 의 `/` 함정이 아예 없다 —
    예전에는 원문자를 토큰 경계로 삼아 그것을 피했다.
    """
    item = row(1, documents=[("1", "Project Charter & R&R"),
                             ("3", "DSEP Readiness & I/O Spec")])
    cell = documents_cell(item)

    assert cell == "1 / 3"
    assert migrate.parse_documents(cell) == [1, 3]


def test_owners_join_on_plus():
    item = row(1, owners=["DSEP 인프라 담당자", "사내 IT·보안"])
    assert owners_cell(item) == "DSEP 인프라 담당자+사내 IT·보안"
    assert migrate.parse_owners(owners_cell(item)) == [
        "DSEP 인프라 담당자", "사내 IT·보안"
    ]


def test_every_status_round_trips_through_the_original_notation():
    """다섯 상태 전부. 하나라도 표기가 어긋나면 재적재 시 조용히 진행전이 된다."""
    for status in ItemStatus:
        label = EXPORT_STATUS[status]
        assert migrate._map_status(label) == status.value, f"{status} → {label!r}"


# =============================================================================
# round-trip 계약 — 이 모듈의 fail-closed 게이트
# =============================================================================
def seed_book():
    return migrate.parse_workbook(SOURCE_XLSX)


@pytest.fixture(scope="module")
def source_book():
    if not SOURCE_XLSX.exists():
        pytest.skip(f"원본 엑셀이 없습니다: {SOURCE_XLSX}")
    return seed_book()


def items_from(book) -> list[ItemOut]:
    """원본 파싱 결과를 그대로 `ItemOut` 으로 되돌린다 (DB 없이 왕복을 만든다)."""
    phase_ids: dict[int, int] = {}
    ms_ids: dict[tuple[int, int], int] = {}
    doc_names = {i: d.name for i, d in enumerate(book.doc_types, start=1)}

    items = []
    for parsed in book.rows:
        phase_ids.setdefault(parsed.phase_seq, len(phase_ids) + 1)
        key = (parsed.phase_seq, parsed.milestone_seq)
        ms_ids.setdefault(key, len(ms_ids) + 1)
        items.append(
            row(
                parsed.no,
                phase_id=phase_ids[parsed.phase_seq],
                phase_no=parsed.phase_seq,
                phase_name=parsed.phase_name,
                milestone_id=ms_ids[key],
                milestone_no=parsed.milestone_seq,
                milestone_name=parsed.milestone_name,
                status=ItemStatus(parsed.status),
                title=parsed.title,
                deliverable=parsed.deliverable,
                documents=[(str(n), doc_names[n]) for n in parsed.doc_numbers],
                owners=list(parsed.owner_names),
            )
        )
    return items


class _Doc:
    """§0.5.10 이후의 문서 — 순서와 이름이 전부다."""

    def __init__(self, parsed, index):
        self.id = index
        self.sort_order = index
        self.name = parsed.name


def test_the_exported_seed_board_reparses_with_parse_workbook(tmp_path, source_book):
    """**계약 그 자체.** 내보낸 파일이 곧 임포트 양식이어야 한다.

    시드 보드(완전 배정 35행)를 내보내고, `db/migrate.py` 의 파서로 다시 읽어
    행 수와 필드가 원본과 일치하는지 본다. 표기를 하나만 흩뜨려도 여기서 깨진다.
    """
    export = BoardExport(
        title="시드 보드",
        items=items_from(source_book),
        documents=[_Doc(d, i) for i, d in enumerate(source_book.doc_types, start=1)],
    )
    path = tmp_path / "roundtrip.xlsx"
    path.write_bytes(build_board_xlsx(export))

    again = migrate.parse_workbook(path)

    assert len(again.rows) == len(source_book.rows) == 35
    assert len(again.doc_types) == len(source_book.doc_types) == 5

    for before, after in zip(source_book.rows, again.rows):
        assert after.no == before.no
        assert after.phase_seq == before.phase_seq
        assert after.phase_name == before.phase_name
        assert after.milestone_seq == before.milestone_seq
        assert after.milestone_name == before.milestone_name
        assert after.title == before.title
        assert after.deliverable == before.deliverable
        assert after.doc_numbers == before.doc_numbers
        assert after.owner_names == before.owner_names
        assert after.status == before.status


def test_the_reparsed_export_also_passes_the_consistency_checks(tmp_path, source_book):
    """파싱만 되는 것과 **적재 가능한 것**은 다르다 — 정합성 검사까지 통과해야 한다."""
    export = BoardExport(
        title="시드 보드",
        items=items_from(source_book),
        documents=[_Doc(d, i) for i, d in enumerate(source_book.doc_types, start=1)],
    )
    path = tmp_path / "consistency.xlsx"
    path.write_bytes(build_board_xlsx(export))

    assert migrate.check_consistency(migrate.parse_workbook(path)) == []


def test_the_round_trip_gate_actually_fails_when_notation_drifts(tmp_path, source_book):
    """게이트가 **실패할 수 있는지** 확인한다 (NEGATIVE CONTROL).

    통과하는 계약 테스트는 없는 것보다 나쁘다. Phase 표기에서 점 하나를 빼고
    같은 파일을 만들면 재파싱이 반드시 죽어야 한다.
    """
    from unittest.mock import patch

    export = BoardExport(
        title="어긋난 표기",
        items=items_from(source_book),
        documents=[_Doc(d, i) for i, d in enumerate(source_book.doc_types, start=1)],
    )
    with patch.object(
        board_xlsx_service, "phase_cell",
        lambda item: f"Phase {item.phase_no} {item.phase_name or ''}".rstrip(),
    ):
        path = tmp_path / "broken.xlsx"
        path.write_bytes(build_board_xlsx(export))

    with pytest.raises(ValueError, match="Phase 형식 불일치"):
        migrate.parse_workbook(path)


# =============================================================================
# 원본 양식과 같은 모양인가
# =============================================================================
def test_the_sheets_match_the_original_layout(source_book):
    """시트 이름·머리말 위치·헤더 행·컬럼 구성을 원본 파일과 직접 대조한다."""
    from openpyxl import load_workbook

    original = load_workbook(SOURCE_XLSX)
    ours = open_xlsx(build_board_xlsx(BoardExport(
        title="양식", items=items_from(source_book),
        documents=[_Doc(d, i) for i, d in enumerate(source_book.doc_types, start=1)],
    )))

    assert ours.sheetnames == original.sheetnames == [BOARD_SHEET, DOC_SHEET]

    board, source = ours[BOARD_SHEET], original[BOARD_SHEET]
    assert [c.value for c in board[BOARD_HEADER_ROW]] == [
        c.value for c in source[BOARD_HEADER_ROW]
    ]
    assert [c.value for c in board[BOARD_HEADER_ROW]] == list(BOARD_HEADERS)
    assert board["A1"].value == source["A1"].value
    assert board.freeze_panes == source.freeze_panes

    docs, doc_source = ours[DOC_SHEET], original[DOC_SHEET]
    # Doc Status 시트는 §0.5.10 에서 순서·문서명 둘로 줄었다 — 원본의 7컬럼과
    # 다른 것이 의도다 (단계/Gate/작성 주체/비고는 전역 문서 모델의 잔재였다).
    assert [c.value for c in docs[DOC_HEADER_ROW] if c.value] == list(DOC_HEADERS)


def test_the_header_is_styled_and_the_data_rows_are_bordered(source_book):
    board = open_xlsx(build_board_xlsx(BoardExport(
        title="서식", items=items_from(source_book),
        documents=[_Doc(d, i) for i, d in enumerate(source_book.doc_types, start=1)],
    )))[BOARD_SHEET]

    head = board.cell(row=BOARD_HEADER_ROW, column=1)
    assert head.font.bold is True
    assert head.fill.fgColor.rgb == "FF1E2761"
    assert head.border.bottom.style == "thin"

    first = board.cell(row=BOARD_FIRST_DATA_ROW, column=1)
    assert first.font.name == "Malgun Gothic"
    assert first.border.left.style == "thin"


def test_phase_rows_are_banded_and_unassigned_rows_are_not():
    """§0.5 팔레트로 밴딩한다. 미배정 행은 무색이어야 구분이 된다."""
    items = [
        row(1, phase_id=1, phase_no=0, phase_name="P0", milestone_id=10, milestone_no=1),
        row(2, phase_id=2, phase_no=1, phase_name="P1", milestone_id=20, milestone_no=1),
        row(3),
    ]
    board = open_xlsx(build_board_xlsx(BoardExport(title="밴딩", items=items)))[BOARD_SHEET]

    p0 = board.cell(row=BOARD_FIRST_DATA_ROW, column=2).fill
    p1 = board.cell(row=BOARD_FIRST_DATA_ROW + 1, column=2).fill
    gray = board.cell(row=BOARD_FIRST_DATA_ROW + 2, column=2).fill

    assert p0.fgColor.rgb != p1.fgColor.rgb, "Phase 마다 다른 색이어야 한다"
    assert gray.fill_type is None, "미배정 행에는 밴딩이 없다"


def test_the_completion_date_is_written_as_a_date():
    items = [row(1, phase_id=1, phase_no=0, phase_name="P", milestone_id=10,
                 milestone_no=1, completion_date=date(2026, 3, 14))]
    board = open_xlsx(build_board_xlsx(BoardExport(title="날짜", items=items)))[BOARD_SHEET]

    cell = board.cell(row=BOARD_FIRST_DATA_ROW, column=9)
    assert cell.value.date() if hasattr(cell.value, "date") else cell.value == date(2026, 3, 14)
    assert cell.number_format == "yyyy-mm-dd"


def test_an_unused_document_is_left_out_of_the_doc_sheet(tmp_path):
    """§0.5.10 팝업 정밀화 — 꺼진 문서는 시트에 싣지 않는다.

    표시 번호가 사용 문서만 세어 매겨지므로, 꺼진 문서를 함께 실으면 시트의 행
    위치와 '관련 문서' 셀의 숫자가 어긋나 재파싱이 깨진다. 화면·전체 현황·PPT
    어디에도 나오지 않는 문서가 내보내기에만 있는 것도 일관성을 해친다.
    """
    class _P:
        def __init__(self, id, name, order, used):
            self.id, self.name, self.sort_order, self.is_used = id, name, order, used

    documents = [_P(1, "쓰는 문서", 1, True), _P(2, "끈 문서", 2, False),
                 _P(3, "쓰는 문서2", 3, True)]
    items = [row(1, phase_id=1, phase_no=0, phase_name="P", milestone_id=10, milestone_no=1,
                 title="행", deliverable="산출물",
                 documents=[("1", "쓰는 문서"), ("2", "쓰는 문서2")])]

    workbook = open_xlsx(build_board_xlsx(
        BoardExport(title="사용만", items=items, documents=documents)
    ))
    docs = workbook[DOC_SHEET]

    names = [docs.cell(row=DOC_HEADER_ROW + 1 + i, column=2).value for i in range(3)]
    assert names == ["쓰는 문서", "쓰는 문서2", None], "꺼진 문서가 시트에 실렸다"
    # 순서 칸도 1..N 으로 촘촘하다 — 시트 위치와 셀 숫자가 어긋나지 않는다.
    assert [docs.cell(row=DOC_HEADER_ROW + 1 + i, column=1).value for i in range(2)] == [1, 2]


def test_a_board_with_an_unused_document_still_round_trips(tmp_path):
    """사용 문서만 실어도 **재파싱이 성립**해야 한다 — 그것이 그 선택의 근거다."""
    class _P:
        def __init__(self, id, name, order, used):
            self.id, self.name, self.sort_order, self.is_used = id, name, order, used

    documents = [_P(1, "쓰는 문서", 1, True), _P(2, "끈 문서", 2, False),
                 _P(3, "쓰는 문서2", 3, True)]
    items = [
        row(no, phase_id=1, phase_no=0, phase_name="P", milestone_id=10, milestone_no=1,
            milestone_name="M", title=f"행 {no}", deliverable=f"산출물 {no}",
            documents=[("1", "쓰는 문서")] if no == 1 else [("2", "쓰는 문서2")],
            owners=["담당"])
        for no in (1, 2)
    ]
    path = tmp_path / "unused.xlsx"
    path.write_bytes(build_board_xlsx(
        BoardExport(title="왕복", items=items, documents=documents)
    ))

    book = migrate.parse_workbook(path)
    assert len(book.doc_types) == 2                      # 사용 문서만
    assert [r.doc_numbers for r in book.rows] == [[1], [2]]
    # 문서 관련 정합성만 본다 — 2행짜리 보드라 DASH_LABELS(35행 상수) 검사는
    # 당연히 걸리고, 그건 이 테스트의 관심사가 아니다.
    problems = [p for p in migrate.check_consistency(book) if "문서" in p]
    assert problems == []


def test_an_empty_board_still_produces_both_sheets():
    """새 프로젝트는 실제로 행이 0개다 — 여기서 터지면 안 된다."""
    workbook = open_xlsx(build_board_xlsx(BoardExport(title="빈 보드", items=[])))

    assert workbook.sheetnames == [BOARD_SHEET, DOC_SHEET]
    board = workbook[BOARD_SHEET]
    assert [c.value for c in board[BOARD_HEADER_ROW]] == list(BOARD_HEADERS)
    assert board.cell(row=BOARD_FIRST_DATA_ROW, column=1).value is None


def test_a_gray_row_board_exports_but_is_not_a_round_trip_candidate(tmp_path):
    """미배정 행이 섞인 보드는 **내보내지되** 재파싱 대상이 아니다.

    파서는 빈 Phase 셀을 형식 오류로 거부한다. 그것이 옳다 — 그 상태는 원본
    양식에 없고, 발행 검증(V1/V2)이 막는 상태이기도 하다.
    """
    items = [
        row(1, phase_id=1, phase_no=0, phase_name="P", milestone_id=10, milestone_no=1,
            title="배정된 행", deliverable="산출물"),
        row(2, title="회색 행"),
    ]
    path = tmp_path / "gray.xlsx"
    path.write_bytes(build_board_xlsx(BoardExport(title="회색", items=items)))

    board = open_xlsx(path.read_bytes())[BOARD_SHEET]
    # openpyxl 은 빈 문자열을 **진짜 빈 셀**로 쓴다 — 되읽으면 None 이다.
    # "빈 셀" 요구사항이 바라는 것이 정확히 이것이다.
    assert board.cell(row=BOARD_FIRST_DATA_ROW + 1, column=2).value is None
    assert board.cell(row=BOARD_FIRST_DATA_ROW + 1, column=3).value is None
    assert board.cell(row=BOARD_FIRST_DATA_ROW + 1, column=4).value == "회색 행"

    with pytest.raises(ValueError):
        migrate.parse_workbook(path)


# =============================================================================
# 엔드포인트
# =============================================================================
@pytest.mark.db
class TestEndpoints:
    @pytest.fixture
    def published(self, db, board):
        for order in (1, 2):
            item = Item(
                version_id=board.published.id, sort_order=order,
                phase_id=board.p0.id, milestone_id=board.m01.id,
                title=f"행 {order}", deliverable=f"산출물 {order}",
            )
            item.owners = [ItemOwner(owner_id=board.o1.id, sort_order=1)]
            item.documents = [ItemDocument(template_document_id=board.d1.id, sort_order=1)]
            db.add(item)
        db.commit()
        return board

    @pytest.fixture
    def project(self, client, published):
        response = client.post(
            f"{API}/projects",
            json={"maker_id": 7, "name": "데모 프로젝트", "template_id": published.wp.id},
        )
        assert response.status_code == 201, response.text
        return response.json()["project"]["id"]

    def test_the_version_export_returns_an_xlsx(self, client, published):
        response = client.get(f"{API}/versions/{published.published.id}/board.xlsx")

        assert response.status_code == 200, response.text
        assert response.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert response.content[:2] == b"PK"

    def test_the_project_export_returns_an_xlsx(self, client, project):
        response = client.get(f"{API}/projects/{project}/board.xlsx")
        assert response.status_code == 200, response.text
        assert response.content[:2] == b"PK"

    def test_the_workbook_carries_the_board_rows(self, client, project):
        board = open_xlsx(
            client.get(f"{API}/projects/{project}/board.xlsx").content
        )[BOARD_SHEET]

        assert board.cell(row=BOARD_FIRST_DATA_ROW, column=1).value == 1
        assert board.cell(row=BOARD_FIRST_DATA_ROW, column=2).value.startswith("Phase 0.")
        assert board.cell(row=BOARD_FIRST_DATA_ROW, column=4).value == "행 1"
        assert board.cell(row=BOARD_FIRST_DATA_ROW, column=5).value == "산출물 1"
        assert board.cell(row=BOARD_FIRST_DATA_ROW, column=6).value == "1"
        assert board.cell(row=BOARD_FIRST_DATA_ROW, column=7).value == "DSEP 인프라 담당자"
        assert board.cell(row=BOARD_FIRST_DATA_ROW, column=8).value == "Not Started"

    def test_the_project_doc_sheet_adds_the_project_columns(self, client, project):
        docs = open_xlsx(
            client.get(f"{API}/projects/{project}/board.xlsx").content
        )[DOC_SHEET]

        headers = [c.value for c in docs[DOC_HEADER_ROW] if c.value]
        assert headers == list(DOC_HEADERS) + ["사용", "문서 링크", "작성 상태"]
        assert docs.cell(row=DOC_HEADER_ROW + 1, column=1).value == 1
        assert docs.cell(row=DOC_HEADER_ROW + 1, column=3).value == "O"
        assert docs.cell(row=DOC_HEADER_ROW + 1, column=5).value == "작성 전"

    def test_the_project_doc_sheet_reflects_saved_settings(self, client, project):
        rows = client.get(f"{API}/projects/{project}/documents").json()["documents"]
        client.put(
            f"{API}/projects/{project}/documents",
            json={"documents": [
                {"id": rows[0]["id"], "name": rows[0]["name"], "is_used": True,
                 "doc_status": "DONE", "link_url": "https://drive.example.com/x"},
                {"id": rows[1]["id"], "name": rows[1]["name"]},
            ]},
        )

        docs = open_xlsx(
            client.get(f"{API}/projects/{project}/board.xlsx").content
        )[DOC_SHEET]
        assert docs.cell(row=DOC_HEADER_ROW + 1, column=4).value == "https://drive.example.com/x"
        assert docs.cell(row=DOC_HEADER_ROW + 1, column=5).value == "완료"

    def test_the_template_doc_sheet_has_no_project_columns(self, client, published):
        docs = open_xlsx(
            client.get(f"{API}/versions/{published.published.id}/board.xlsx").content
        )[DOC_SHEET]

        headers = [c.value for c in docs[DOC_HEADER_ROW] if c.value]
        assert headers == list(DOC_HEADERS)

    def test_the_filename_header_uses_rfc_5987(self, client, project):
        disposition = client.get(
            f"{API}/projects/{project}/board.xlsx"
        ).headers["content-disposition"]

        assert disposition.startswith("attachment;")
        assert f"filename*=UTF-8''{quote('데모 프로젝트.xlsx', safe='')}" in disposition
        assert re.search(r'filename="[\x20-\x7e]*"', disposition)

    def test_the_version_filename_names_the_template_and_number(self, client, published):
        disposition = client.get(
            f"{API}/versions/{published.published.id}/board.xlsx"
        ).headers["content-disposition"]
        assert quote("테스트 보드 v1.xlsx", safe="") in disposition

    def test_a_project_with_no_rows_exports_without_erroring(self, client, board):
        project_id = client.post(
            f"{API}/projects",
            json={"maker_id": 7, "name": "빈 프로젝트", "template_id": board.wp.id},
        ).json()["project"]["id"]

        response = client.get(f"{API}/projects/{project_id}/board.xlsx")
        assert response.status_code == 200, response.text
        assert open_xlsx(response.content).sheetnames == [BOARD_SHEET, DOC_SHEET]

    def test_a_gray_row_exports_as_blank_cells(self, client, project):
        assert client.post(f"{API}/projects/{project}/items").status_code == 201

        board = open_xlsx(
            client.get(f"{API}/projects/{project}/board.xlsx").content
        )[BOARD_SHEET]
        last = BOARD_FIRST_DATA_ROW + 2
        assert board.cell(row=last, column=1).value == 3
        assert board.cell(row=last, column=2).value is None
        assert board.cell(row=last, column=3).value is None

    def test_an_archived_version_can_still_be_exported(self, db, client, published):
        """불변 규칙은 **쓰기**를 막는 것이지 읽기를 막는 것이 아니다."""
        from app.models import VersionStatus

        published.published.status = VersionStatus.ARCHIVED
        db.commit()

        response = client.get(f"{API}/versions/{published.published.id}/board.xlsx")
        assert response.status_code == 200, response.text

    def test_unknown_ids_are_404(self, client):
        assert client.get(f"{API}/projects/999999/board.xlsx").status_code == 404
        assert client.get(f"{API}/versions/999999/board.xlsx").status_code == 404

    def test_both_exports_are_read_only_routes(self, client):
        paths = client.app.openapi()["paths"]
        assert set(paths["/api/v1/projects/{project_id}/board.xlsx"]) == {"get"}
        assert set(paths["/api/v1/versions/{version_id}/board.xlsx"]) == {"get"}
