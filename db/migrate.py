"""엑셀 → `iai-test` DB 임포트 스크립트.

`docs/Work Package.xlsx` 의 두 시트를 읽어

  * **Project Board** (35행) → `wp_phases` / `wp_milestones` / `wp_owners` /
    `wp_items` / `wp_item_documents` / `wp_item_owners`
  * **Doc Status** (5행)     → `wp_template_documents` (템플릿 스코프, §0.5.10)

로 적재하고, 적재 결과를 v1 `PUBLISHED` 버전으로 등록한다.

파싱 규칙 (plan.md §1.1 / CLAUDE.md "Working with the source documents")

  ``Phase``      ``Phase 0. Pre-Infrastructure Setup``
                 → ``phases(seq_no=0, name='Pre-Infrastructure Setup')``
  ``Milestone``  ``0.1 DSEP 환경 Gap 및 자원 구성``
                 → ``milestones(seq_no=1, name='DSEP 환경 Gap 및 자원 구성')``
                   앞자리 0 은 **저장하지 않고** 소속 Phase 에서 파생한다.
  ``관련 문서``   ``/`` 구분 다중값 → ``wp_item_documents`` (N:M)
  ``Owner``      ``+`` 구분 다중값 → ``wp_item_owners`` (N:M)

사용법::

    python db/migrate.py                      # 스키마 적용 + 엑셀 적재
    python db/migrate.py --emit-sql           # 위 + db/seed.sql 재생성
    python db/migrate.py --dry-run            # DB 를 건드리지 않고 파싱 결과만 검사
    python db/migrate.py --with-dev-seed      # dev_seed.sql(wp_dev_makers) 도 적용

주의: 콘솔이 cp949 라 한글 출력이 깨진다. 이 스크립트는 요약 통계(숫자)만
출력하고, 파싱 결과 원문이 필요하면 ``--dump`` 로 UTF-8 JSON 파일에 쓴다.
"""

from __future__ import annotations

import argparse
import io
import json
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

import pymysql
from openpyxl import load_workbook

# `backend/.env` (`backend/.env.example` 참고). 아래 상수를 읽기 **전에** 올려야 한다.
#
# 이 스크립트는 `db/` 에 있지만 `.env` 는 `backend/` 에 둔다 — 접속 정보 파일이
# 저장소에 하나뿐이어야 두 벌이 갈릴 일이 없고, 사용자가 서버를 띄우는 자리가
# 거기이기 때문이다.
#
# python-dotenv 는 requirements-dev.txt 에만 있으므로, 없으면 건너뛰고 셸
# 환경변수만 쓴다 — requirements.txt 만 설치한 사람도 이 스크립트는 돌아야 한다.
# 기존 환경변수는 덮어쓰지 않는다.
try:  # pragma: no cover - 설치 여부에 따른 분기
    from dotenv import load_dotenv

    load_dotenv(Path(__file__).resolve().parent.parent / "backend" / ".env")
except ImportError:  # pragma: no cover
    pass

# --- 접속 정보 -----------------------------------------------------------------
# **환경변수로 받는다** — 저장소에 비밀번호를 두지 않는다. 비밀번호만 기본값이
# 없고, 없으면 `connect()` 가 무엇을 설정해야 하는지 말하며 멈춘다.
# 백엔드 애플리케이션은 같은 이유로 WP_DB_DSN 을 쓴다 (pydantic-settings).
#
# `db/verify.py` 가 이 상수들을 그대로 import 해 쓴다.
DB_HOST = os.environ.get("WP_DB_HOST", "localhost")
DB_PORT = int(os.environ.get("WP_DB_PORT", "3306"))
DB_USER = os.environ.get("WP_DB_USER", "user01")
DB_PASSWORD = os.environ.get("WP_DB_PASSWORD", "")
DB_NAME = os.environ.get("WP_DB_NAME", "iai-test")  # 하이픈 → raw SQL 에서 항상 백틱

REPO_ROOT = Path(__file__).resolve().parent.parent
EXCEL_PATH = REPO_ROOT / "docs" / "Work Package.xlsx"
SCHEMA_PATH = REPO_ROOT / "db" / "schema.sql"
DEV_SEED_PATH = REPO_ROOT / "db" / "dev_seed.sql"
SEED_OUT_PATH = REPO_ROOT / "db" / "seed.sql"

# 엑셀 35행은 **템플릿 v1 PUBLISHED** 로 들어간다 (plan.md §0.1).
# 설비사별 인스턴스(프로젝트)는 여기서 만들지 않는다 — 데모 프로젝트는
# dev_seed.sql 에만 있다. 기준 데이터에 특정 설비사의 흔적을 남기지 않기 위해서다.
TEMPLATE = {
    "code": "DSEP-AI-BOARD",
    "name": "DSEP AI Project Board",
    "description": "docs/Work Package.xlsx 의 Project Board 시트를 웹으로 이관한 초기 보드",
    "phase_start_no": 0,
}

# 대시보드 카드 라벨 (plan.md §0.5-1). 행 No → 라벨.
#
# **엑셀에서 읽지 않는다.** 출처는 `docs/dashboard.jpg` 이고, 엑셀의 `Key Action
# Item` 은 카드에 싣기엔 너무 긴 서술문이라 요약이 따로 필요하다. 엑셀 양식은
# 고정이므로 행 No 로 못박아도 어긋날 여지가 없고, 실제로 어긋나면
# `check_consistency` 가 잡는다.
#
# 여기 있는 것은 **초기값**일 뿐이다. 이후에는 그리드의 "대시보드 표시" 컬럼이
# 정본이며, 이 상수를 고쳐도 이미 적재된 DB 는 바뀌지 않는다.
DASH_LABELS = {
    1: "Gap·자원 계획",
    2: "Workspace 구성",
    3: "사내 자원 연결",
    4: "E2E 검증 · G0",
    5: "범위·문제 정의",
    6: "Use Case·I/O",
    7: "KPI·합격 기준",
    8: "RACI 책임 경계",
    9: "소통·승인 체계",
    10: "협약 체결",
    11: "통합 계획",
    12: "데이터 제공 계획",
    13: "I/O 규격",
    14: "Kick-off · G1",
    15: "개발계획 접수",
    16: "통합 일정 정렬",
    17: "Input 제공 추적",
    18: "인프라 지원",
    19: "진행현황 점검",
    20: "RAID·Action",
    21: "결정·변경 관리",
    22: "중간 Review",
    23: "Candidate 제출",
    24: "평가 준비 점검",
    25: "평가 인계 · G2",
    26: "평가 계획",
    27: "평가결과 종합",
    28: "판정 · G3",
    29: "Pilot 계획",
    30: "Pilot 결과 종합",
    31: "Go/No-Go · G4",
    32: "최종 인계",
    33: "종료 처리",
    34: "표준 개선",
    35: "확대 승인 · G5",
}

#: `wp_items.dash_label` / `wp_project_items.dash_label` 의 컬럼 폭.
DASH_LABEL_MAX = 60

STATUS_MAP = {
    "not started": "NOT_STARTED",
    "in progress": "IN_PROGRESS",
    "done": "DONE",
    "hold": "HOLD",
    "n/a": "NA",
    "na": "NA",
}

PHASE_RE = re.compile(r"^\s*Phase\s+(\d+)\s*[.．]\s*(.+?)\s*$", re.IGNORECASE)
MILESTONE_RE = re.compile(r"^\s*(\d+)\.(\d+)\s+(.+?)\s*$")
DOC_CODE_RE = re.compile(r"^\s*([①②③④⑤⑥⑦⑧⑨⑩])\s*(.*?)\s*$")
# `관련 문서` 는 `/` 구분 다중값이지만, 문서명 자체에 `/` 가 들어간다
# (`② DSEP Readiness & I/O Spec`). 따라서 `/` 로 split 하면 안 되고,
# **원문자(①~⑩)를 토큰 시작점**으로 삼아 잘라야 한다.
DOC_TOKEN_RE = re.compile(r"([①②③④⑤⑥⑦⑧⑨⑩])\s*([^①②③④⑤⑥⑦⑧⑨⑩]*)")


# =============================================================================
# 파싱
# =============================================================================
@dataclass
class ParsedRow:
    no: int
    phase_seq: int
    phase_name: str
    milestone_seq: int
    milestone_name: str
    title: str
    deliverable: str
    doc_numbers: list[int]
    owner_names: list[str]
    status: str
    completion_date: object = None
    #: 대시보드 카드 라벨. 엑셀이 아니라 `DASH_LABELS` 에서 온다.
    dash_label: str | None = None


@dataclass
class ParsedDocType:
    """Doc Status 시트의 한 줄.

    §0.5.10 이후 문서가 갖는 것은 **표시 번호와 이름**뿐이다. 원본 엑셀의
    단계/Gate/작성 주체/비고는 전역 문서 모델의 잔재였고 저장되지 않는다 —
    파싱은 하되 버린다 (원본 파일을 계속 읽을 수 있어야 하므로).
    """

    #: 원본의 첫 칸 (`①` 또는 `1`). 표시 번호는 `sort_order` 다.
    code: str
    name: str
    sort_order: int


@dataclass
class ParsedBook:
    rows: list[ParsedRow] = field(default_factory=list)
    doc_types: list[ParsedDocType] = field(default_factory=list)


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _map_status(value: str) -> str:
    return STATUS_MAP.get(value.strip().lower(), "NOT_STARTED")


def _find_header(rows: list[list], *first_col_labels: str) -> int:
    """헤더 행을 찾는다. **여러 라벨을 받는다.**

    Doc Status 시트의 첫 컬럼이 원본 엑셀에서는 `No`, 내보내기(§0.5.10)에서는
    `순서` 다. 문서 표기를 숫자로 통일하면서 라벨도 바뀌었고, 두 파일을 모두
    읽어야 하므로 둘 다 받는다 — `parse_documents` 가 두 표기를 받는 것과 같은
    이유다.
    """
    for idx, row in enumerate(rows):
        if row and _text(row[0]) in first_col_labels:
            return idx
    raise ValueError(f"헤더 행(첫 컬럼 {' 또는 '.join(first_col_labels)})을 찾지 못했습니다.")


#: 원문자 → 표시 번호. 원본 엑셀은 ①②③ 을 쓰고 §0.5.10 이후 표기는 숫자다.
CIRCLED_TO_NUMBER = {c: i for i, c in enumerate("①②③④⑤⑥⑦⑧⑨⑩", start=1)}


def parse_documents(cell: str) -> list[int]:
    """`관련 문서` 셀에서 **문서 표시 번호** 목록을 뽑는다 (plan.md §0.5.10 ③).

    두 표기를 모두 받는다.

    * **숫자** (`1 / 3`) — 내보내기가 쓰는 현재 표기. round-trip 대상이다.
    * **원문자** (`① Project Charter / ② …`) — 원본 `docs/Work Package.xlsx` 를
      최초 임포트할 때만 나온다. ①=1 로 대응시킨다.

    원문자 쪽은 문서명 자체에 `/` 가 들어 있어(`DSEP Readiness & I/O Spec`) 단순
    split 이 깨지므로 **원문자를 토큰 경계**로 삼는다. 숫자 표기에는 이름이 없어
    그 함정이 아예 없다 — 표기를 숫자로 통일한 이유 중 하나다.
    """
    cell = cell.strip()
    if not cell:
        return []

    if DOC_CODE_RE.match(cell):
        numbers: list[int] = []
        for code, _name in DOC_TOKEN_RE.findall(cell):
            value = CIRCLED_TO_NUMBER[code]
            if value not in numbers:
                numbers.append(value)
        return numbers

    numbers = []
    for token in cell.split("/"):
        token = token.strip()
        if not token:
            continue
        if not token.isdigit():
            raise ValueError(
                f"'관련 문서' 셀을 읽을 수 없습니다 (숫자 또는 원문자여야 합니다): {cell!r}"
            )
        value = int(token)
        if value not in numbers:
            numbers.append(value)
    return numbers


def parse_owners(cell: str) -> list[str]:
    """`Owner` 셀을 `+` 로 분리해 담당 주체 목록을 돌려준다."""
    names: list[str] = []
    for token in cell.split("+"):
        token = token.strip()
        if token and token not in names:
            names.append(token)
    return names


def parse_workbook(path: Path) -> ParsedBook:
    wb = load_workbook(path, data_only=True, read_only=True)
    book = ParsedBook()

    # --- Project Board ---------------------------------------------------
    board = [list(r) for r in wb["Project Board"].iter_rows(values_only=True)]
    header = _find_header(board, "No")
    for raw in board[header + 1 :]:
        if not raw or _text(raw[0]) == "":
            continue
        phase_cell, milestone_cell = _text(raw[1]), _text(raw[2])

        pm = PHASE_RE.match(phase_cell)
        if not pm:
            raise ValueError(f"Phase 형식 불일치: {phase_cell!r}")
        mm = MILESTONE_RE.match(milestone_cell)
        if not mm:
            raise ValueError(f"Milestone 형식 불일치: {milestone_cell!r}")

        phase_seq = int(pm.group(1))
        major, minor = int(mm.group(1)), int(mm.group(2))
        # Milestone 앞자리는 소속 Phase 번호와 반드시 일치해야 한다.
        # (일치하므로 저장하지 않고 파생값으로 쓸 수 있다 — plan.md §2.1)
        if major != phase_seq:
            raise ValueError(
                f"{raw[0]}행: Milestone 앞자리({major})가 Phase 번호({phase_seq})와 다릅니다."
            )

        book.rows.append(
            ParsedRow(
                no=int(raw[0]),
                phase_seq=phase_seq,
                phase_name=pm.group(2),
                milestone_seq=minor,
                milestone_name=mm.group(3),
                title=_text(raw[3]),
                deliverable=_text(raw[4]),
                doc_numbers=parse_documents(_text(raw[5])),
                owner_names=parse_owners(_text(raw[6])),
                status=_map_status(_text(raw[7])),
                completion_date=raw[8] if len(raw) > 8 else None,
                dash_label=DASH_LABELS.get(int(raw[0])),
            )
        )

    # --- Doc Status ------------------------------------------------------
    docs = [list(r) for r in wb["Doc Status"].iter_rows(values_only=True)]
    header = _find_header(docs, "No", "순서")
    for order, raw in enumerate(docs[header + 1 :], start=1):
        if not raw or _text(raw[0]) == "":
            continue
        # 원본 엑셀은 7컬럼, 내보내기는 2컬럼이다. 앞의 둘만 읽으면 둘 다 통한다.
        book.doc_types.append(
            ParsedDocType(code=_text(raw[0]), name=_text(raw[1]), sort_order=order)
        )

    wb.close()
    return book


def check_consistency(book: ParsedBook) -> list[str]:
    """적재 전에 원본이 설계 전제를 만족하는지 확인한다."""
    problems: list[str] = []

    # 1) No 는 1..N 연속
    expected = list(range(1, len(book.rows) + 1))
    if [r.no for r in book.rows] != expected:
        problems.append("No 컬럼이 1..N 으로 연속하지 않습니다.")

    # 2) Phase 블록 연속성 — 같은 Phase 의 행은 붙어 있어야 한다 (plan.md §2.2)
    seen_phase: set[int] = set()
    prev_phase = None
    for row in book.rows:
        if row.phase_seq != prev_phase:
            if row.phase_seq in seen_phase:
                problems.append(f"{row.no}행: Phase {row.phase_seq} 블록이 연속되지 않습니다.")
            seen_phase.add(row.phase_seq)
            prev_phase = row.phase_seq

    # 3) Milestone 블록 연속성 (Phase 내부)
    seen_ms: set[tuple[int, int]] = set()
    prev_key = None
    for row in book.rows:
        key = (row.phase_seq, row.milestone_seq)
        if key != prev_key:
            if key in seen_ms:
                problems.append(
                    f"{row.no}행: Milestone {key[0]}.{key[1]} 블록이 연속되지 않습니다."
                )
            seen_ms.add(key)
            prev_key = key

    # 4) 같은 Phase 번호에 이름이 두 개 이상 붙어 있지 않은지
    names: dict[int, str] = {}
    for row in book.rows:
        if names.setdefault(row.phase_seq, row.phase_name) != row.phase_name:
            problems.append(f"Phase {row.phase_seq} 의 이름이 행마다 다릅니다.")

    # 5) 참조된 문서 번호가 Doc Status 에 모두 존재하는지
    known = set(range(1, len(book.doc_types) + 1))
    for row in book.rows:
        for number in row.doc_numbers:
            if number not in known:
                problems.append(f"{row.no}행: Doc Status 에 없는 문서 번호 {number}")

    # 6) 대시보드 라벨은 엑셀이 아니라 행 No 로 붙는다. 엑셀 행 수가 바뀌면
    #    라벨이 엉뚱한 행에 붙거나 조용히 비게 되므로, 양쪽 집합이 정확히
    #    일치하는지 확인한다.
    row_numbers = {r.no for r in book.rows}
    stray = sorted(set(DASH_LABELS) - row_numbers)
    if stray:
        problems.append(f"DASH_LABELS 에 엑셀에 없는 행 No 가 있습니다: {stray}")
    unlabeled = sorted(row_numbers - set(DASH_LABELS))
    if unlabeled:
        problems.append(f"대시보드 라벨이 없는 행 No: {unlabeled}")
    for no, label in sorted(DASH_LABELS.items()):
        if len(label) > DASH_LABEL_MAX:
            problems.append(f"{no}행: 대시보드 라벨이 {DASH_LABEL_MAX}자를 넘습니다 ({len(label)}자)")

    return problems


# =============================================================================
# 적재
# =============================================================================
def connect(database: str | None = None) -> pymysql.connections.Connection:
    # 비밀번호가 없으면 "Access denied" 로 끝나서 원인이 안 보인다. 먼저 말한다.
    if not DB_PASSWORD:
        sys.exit(
            "WP_DB_PASSWORD 환경변수가 필요합니다.\n"
            "  PowerShell: $env:WP_DB_PASSWORD = 'xxxx'\n"
            "  bash      : export WP_DB_PASSWORD=xxxx"
        )
    return pymysql.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        database=database,
        charset="utf8mb4",
        autocommit=False,
    )


def split_statements(sql: str) -> list[str]:
    """세미콜론 기준 분리. 이 저장소의 DDL 은 트리거/프로시저를 쓰지 않는다."""
    out, buf, in_line_comment = [], [], False
    for line in sql.splitlines():
        stripped = line.strip()
        if stripped.startswith("--") or not stripped:
            continue
        buf.append(line)
        if stripped.endswith(";"):
            statement = "\n".join(buf).strip().rstrip(";").strip()
            if statement:
                out.append(statement)
            buf = []
    tail = "\n".join(buf).strip()
    if tail:
        out.append(tail)
    return out


def apply_sql_file(path: Path, database: str | None) -> int:
    conn = connect(database)
    count = 0
    try:
        with conn.cursor() as cur:
            for stmt in split_statements(path.read_text(encoding="utf-8")):
                cur.execute(stmt)
                count += 1
        conn.commit()
    finally:
        conn.close()
    return count


def load(book: ParsedBook) -> dict:
    """파싱 결과를 DB 에 적재하고 요약 통계를 돌려준다 (전체 1 트랜잭션)."""
    conn = connect(DB_NAME)
    stats: dict[str, int] = {}
    try:
        with conn.cursor() as cur:
            # --- 재실행 가능하도록 기존 데이터 제거 (FK 역순) ------------------
            cur.execute("SELECT id FROM `wp_templates` WHERE code=%s", (TEMPLATE["code"],))
            existing = cur.fetchone()
            if existing:
                wp_id = existing[0]
                cur.execute(
                    "DELETE wio FROM `wp_item_owners` wio "
                    "JOIN `wp_items` i ON i.id = wio.item_id "
                    "JOIN `wp_versions` v ON v.id = i.version_id WHERE v.template_id=%s", (wp_id,))
                cur.execute(
                    "DELETE wid FROM `wp_item_documents` wid "
                    "JOIN `wp_items` i ON i.id = wid.item_id "
                    "JOIN `wp_versions` v ON v.id = i.version_id WHERE v.template_id=%s", (wp_id,))
                cur.execute(
                    "DELETE i FROM `wp_items` i JOIN `wp_versions` v ON v.id = i.version_id "
                    "WHERE v.template_id=%s", (wp_id,))
                cur.execute("UPDATE `wp_versions` SET source_version_id=NULL WHERE template_id=%s", (wp_id,))
                cur.execute("DELETE FROM `wp_versions` WHERE template_id=%s", (wp_id,))
                cur.execute("DELETE FROM `wp_milestones` WHERE template_id=%s", (wp_id,))
                cur.execute("DELETE FROM `wp_phases` WHERE template_id=%s", (wp_id,))
                cur.execute("DELETE FROM `wp_owners` WHERE template_id=%s", (wp_id,))
                cur.execute("DELETE FROM `wp_template_documents` WHERE template_id=%s", (wp_id,))
                cur.execute("DELETE FROM `wp_templates` WHERE id=%s", (wp_id,))

            # --- 템플릿 --------------------------------------------------------
            cur.execute(
                "INSERT INTO `wp_templates` (code, name, description, phase_start_no, is_active) "
                "VALUES (%s,%s,%s,%s,1)",
                (TEMPLATE["code"], TEMPLATE["name"],
                 TEMPLATE["description"], TEMPLATE["phase_start_no"]),
            )
            wp_id = cur.lastrowid

            # --- 문서 (템플릿 스코프 — §0.5.10) ---------------------------------
            # 표시 번호는 Doc Status 시트의 순서 그대로 1..N.
            doc_id_by_number: dict[int, int] = {}
            for order, d in enumerate(book.doc_types, start=1):
                cur.execute(
                    "INSERT INTO `wp_template_documents` (template_id, name, sort_order, is_active) "
                    "VALUES (%s,%s,%s,1)",
                    (wp_id, d.name, order),
                )
                doc_id_by_number[order] = cur.lastrowid
            stats["documents"] = len(doc_id_by_number)

            # --- Phase / Milestone (행 등장 순서 = 번호 순서) --------------------
            phase_ids: dict[int, int] = {}
            for row in book.rows:
                if row.phase_seq in phase_ids:
                    continue
                cur.execute(
                    "INSERT INTO `wp_phases` (template_id, name, seq_no, is_active) VALUES (%s,%s,%s,1)",
                    (wp_id, row.phase_name, row.phase_seq),
                )
                phase_ids[row.phase_seq] = cur.lastrowid
            stats["phases"] = len(phase_ids)

            ms_ids: dict[tuple[int, int], int] = {}
            for row in book.rows:
                key = (row.phase_seq, row.milestone_seq)
                if key in ms_ids:
                    continue
                cur.execute(
                    "INSERT INTO `wp_milestones` (template_id, phase_id, name, seq_no, is_active) "
                    "VALUES (%s,%s,%s,%s,1)",
                    (wp_id, phase_ids[row.phase_seq], row.milestone_name, row.milestone_seq),
                )
                ms_ids[key] = cur.lastrowid
            stats["milestones"] = len(ms_ids)

            # --- Owner 기준정보 (WP 스코프) --------------------------------------
            owner_ids: dict[str, int] = {}
            for row in book.rows:
                for name in row.owner_names:
                    if name in owner_ids:
                        continue
                    cur.execute(
                        "INSERT INTO `wp_owners` (template_id, name, sort_order, is_active) "
                        "VALUES (%s,%s,%s,1)",
                        (wp_id, name, len(owner_ids) + 1),
                    )
                    owner_ids[name] = cur.lastrowid
            stats["owners"] = len(owner_ids)

            # --- v1 PUBLISHED 버전 ----------------------------------------------
            cur.execute(
                "INSERT INTO `wp_versions` (template_id, version_number, status, notes, "
                "phase_start_no, published_at, created_by, published_by) "
                "VALUES (%s, 1, 'PUBLISHED', %s, %s, NOW(), 'migrate.py', 'migrate.py')",
                # 발행 버전은 표시 파라미터를 스스로 들고 있어야 한다 (plan.md §2.4).
                (wp_id, "docs/Work Package.xlsx 최초 임포트", TEMPLATE["phase_start_no"]),
            )
            version_id = cur.lastrowid

            # --- 행 + N:M ---------------------------------------------------------
            doc_links = owner_links = 0
            for order, row in enumerate(book.rows, start=1):
                cur.execute(
                    "INSERT INTO `wp_items` (version_id, sort_order, phase_id, milestone_id, title, "
                    "deliverable, dash_label, gate_code, status, completion_date, origin) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,NULL,%s,%s,'INHERITED')",
                    (version_id, order, phase_ids[row.phase_seq],
                     ms_ids[(row.phase_seq, row.milestone_seq)], row.title,
                     row.deliverable, row.dash_label, row.status, row.completion_date),
                )
                item_id = cur.lastrowid
                for i, number in enumerate(row.doc_numbers, start=1):
                    cur.execute(
                        "INSERT INTO `wp_item_documents` (item_id, template_document_id, sort_order) "
                        "VALUES (%s,%s,%s)",
                        (item_id, doc_id_by_number[number], i),
                    )
                    doc_links += 1
                for i, name in enumerate(row.owner_names, start=1):
                    cur.execute(
                        "INSERT INTO `wp_item_owners` (item_id, owner_id, sort_order) VALUES (%s,%s,%s)",
                        (item_id, owner_ids[name], i),
                    )
                    owner_links += 1

            stats.update(
                template_id=wp_id,
                version_id=version_id,
                items=len(book.rows),
                item_documents=doc_links,
                item_owners=owner_links,
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return stats


# =============================================================================
# seed.sql 생성
# =============================================================================
def _q(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    text = str(value).replace("\\", "\\\\").replace("'", "''")
    return f"'{text}'"


def emit_seed_sql(book: ParsedBook, path: Path) -> None:
    """파싱 결과를 결정적(deterministic)인 INSERT 문으로 직렬화한다.

    id 를 명시해 시드가 재현 가능하도록 만든다. migrate.py 로 적재한 DB 와
    동일한 상태를 SQL 만으로 재구성할 수 있어야 하기 때문이다.
    """
    doc_id = {i: i for i, _d in enumerate(book.doc_types, start=1)}
    phase_id: dict[int, int] = {}
    phase_name: dict[int, str] = {}
    for r in book.rows:
        if r.phase_seq not in phase_id:
            phase_id[r.phase_seq] = len(phase_id) + 1
            phase_name[r.phase_seq] = r.phase_name
    ms_id: dict[tuple[int, int], int] = {}
    ms_name: dict[tuple[int, int], str] = {}
    for r in book.rows:
        key = (r.phase_seq, r.milestone_seq)
        if key not in ms_id:
            ms_id[key] = len(ms_id) + 1
            ms_name[key] = r.milestone_name
    owner_id: dict[str, int] = {}
    for r in book.rows:
        for n in r.owner_names:
            owner_id.setdefault(n, len(owner_id) + 1)

    L: list[str] = []
    add = L.append
    add("-- =============================================================================")
    add("-- 초기 시드 — docs/Work Package.xlsx 원본 데이터")
    add("--")
    add("-- **자동 생성 파일이다. 직접 편집하지 말고 `python db/migrate.py --emit-sql` 로")
    add("-- 재생성한다.** (생성기: db/migrate.py)")
    add("--")
    add("-- 내용: 템플릿 1건 / v1 PUBLISHED 버전 1건 /")
    add(f"--       Phase {len(phase_id)}건 / Milestone {len(ms_id)}건 / Owner {len(owner_id)}건 /")
    add(f"--       문서 {len(doc_id)}건 / 행 {len(book.rows)}건")
    add("--")
    add("-- 전제: db/schema.sql 이 먼저 적용되어 있을 것.")
    add("-- 템플릿은 **중앙 기준 데이터**라 maker 개념이 없다 (plan.md §0.1).")
    add("-- 설비사별 인스턴스(프로젝트)는 이 파일에 없다 — db/dev_seed.sql 참고.")
    add("-- =============================================================================")
    add("")
    add("USE `iai-test`;")
    add("")
    add("SET NAMES utf8mb4;")
    add("")
    add("-- 재실행 가능하도록 기존 시드 제거 (FK 역순)")
    add("DELETE wio FROM `wp_item_owners` wio JOIN `wp_items` i ON i.id=wio.item_id")
    add("  JOIN `wp_versions` v ON v.id=i.version_id WHERE v.template_id=1;")
    add("DELETE wid FROM `wp_item_documents` wid JOIN `wp_items` i ON i.id=wid.item_id")
    add("  JOIN `wp_versions` v ON v.id=i.version_id WHERE v.template_id=1;")
    add("DELETE i FROM `wp_items` i JOIN `wp_versions` v ON v.id=i.version_id WHERE v.template_id=1;")
    add("DELETE FROM `wp_versions`      WHERE template_id=1;")
    add("DELETE FROM `wp_milestones`    WHERE template_id=1;")
    add("DELETE FROM `wp_phases`        WHERE template_id=1;")
    add("DELETE FROM `wp_owners`        WHERE template_id=1;")
    add("DELETE FROM `wp_template_documents` WHERE template_id=1;")
    add("DELETE FROM `wp_templates` WHERE id=1;")
    add("")

    add("-- --- 템플릿 --------------------------------------------------------------------")
    add("INSERT INTO `wp_templates` (id, code, name, description, phase_start_no, is_active)")
    add("VALUES (1, {}, {}, {}, {}, 1);".format(
        _q(TEMPLATE["code"]), _q(TEMPLATE["name"]),
        _q(TEMPLATE["description"]), TEMPLATE["phase_start_no"]))
    add("")

    add("-- --- 문서 (템플릿 스코프 — plan.md §0.5.10) ---------------------------------")
    add("INSERT INTO `wp_template_documents` (id, template_id, name, sort_order, is_active) VALUES")
    add(",\n".join(
        "  ({}, 1, {}, {}, 1)".format(i, _q(d.name), i)
        for i, d in enumerate(book.doc_types, start=1)) + ";")
    add("")

    add("-- --- Phase (번호를 이름에서 분리 — plan.md §2.1) --------------------------------")
    add("INSERT INTO `wp_phases` (id, template_id, name, seq_no, is_active) VALUES")
    add(",\n".join(
        "  ({}, 1, {}, {}, 1)".format(phase_id[s], _q(phase_name[s]), s)
        for s in sorted(phase_id)) + ";")
    add("")

    add("-- --- Milestone (seq_no 는 뒷자리만. 앞자리는 Phase 에서 파생) --------------------")
    add("INSERT INTO `wp_milestones` (id, template_id, phase_id, name, seq_no, is_active) VALUES")
    add(",\n".join(
        "  ({}, 1, {}, {}, {}, 1)".format(ms_id[k], phase_id[k[0]], _q(ms_name[k]), k[1])
        for k in sorted(ms_id)) + ";")
    add("")

    add("-- --- Owner 기준정보 (`Owner` 컬럼의 `+` 분리 결과) -------------------------------")
    add("INSERT INTO `wp_owners` (id, template_id, name, sort_order, is_active) VALUES")
    add(",\n".join(
        "  ({}, 1, {}, {}, 1)".format(oid, _q(name), oid)
        for name, oid in sorted(owner_id.items(), key=lambda kv: kv[1])) + ";")
    add("")

    add("-- --- v1 PUBLISHED 버전 ---------------------------------------------------------")
    add("INSERT INTO `wp_versions`")
    add("  (id, template_id, version_number, status, notes, phase_start_no,")
    add("   published_at, created_by, published_by)")
    add("VALUES (1, 1, 1, 'PUBLISHED', 'docs/Work Package.xlsx 최초 임포트', {},"
        .format(TEMPLATE["phase_start_no"]))
    add("        NOW(), 'seed.sql', 'seed.sql');")
    add("")

    add("-- --- 행 35건 --------------------------------------------------------------------")
    add("INSERT INTO `wp_items`")
    add("  (id, version_id, sort_order, phase_id, milestone_id, title, deliverable, dash_label,")
    add("   gate_code, status, completion_date, origin)")
    add("VALUES")
    add(",\n".join(
        "  ({}, 1, {}, {}, {}, {}, {}, {}, NULL, {}, {}, 'INHERITED')".format(
            r.no, r.no, phase_id[r.phase_seq], ms_id[(r.phase_seq, r.milestone_seq)],
            _q(r.title), _q(r.deliverable), _q(r.dash_label), _q(r.status),
            _q(r.completion_date.date().isoformat() if hasattr(r.completion_date, "date") else r.completion_date))
        for r in book.rows) + ";")
    add("")

    add("-- --- 행 ↔ 문서 (N:M) ------------------------------------------------------------")
    add("INSERT INTO `wp_item_documents` (item_id, template_document_id, sort_order) VALUES")
    add(",\n".join(
        "  ({}, {}, {})".format(r.no, doc_id[n], i)
        for r in book.rows for i, n in enumerate(r.doc_numbers, start=1)) + ";")
    add("")

    add("-- --- 행 ↔ Owner (N:M) -----------------------------------------------------------")
    add("INSERT INTO `wp_item_owners` (item_id, owner_id, sort_order) VALUES")
    add(",\n".join(
        "  ({}, {}, {})".format(r.no, owner_id[n], i)
        for r in book.rows for i, n in enumerate(r.owner_names, start=1)) + ";")
    add("")

    path.write_text("\n".join(L), encoding="utf-8")


# =============================================================================
# CLI
# =============================================================================
def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Work Package 엑셀 → iai-test DB 임포트")
    ap.add_argument("--dry-run", action="store_true", help="DB 를 건드리지 않고 파싱/검사만 수행")
    ap.add_argument("--emit-sql", action="store_true", help="db/seed.sql 재생성")
    ap.add_argument("--with-dev-seed", action="store_true", help="dev_seed.sql (wp_dev_makers) 도 적용")
    ap.add_argument("--skip-schema", action="store_true", help="schema.sql 적용 건너뛰기")
    ap.add_argument(
        "--apply-migrations",
        action="store_true",
        help="db/migrations/ 를 번호순 적용하고 종료 (데이터는 건드리지 않음)",
    )
    ap.add_argument("--dump", metavar="PATH", help="파싱 결과를 UTF-8 JSON 으로 기록")
    args = ap.parse_args(argv)

    if args.apply_migrations:
        # 개발 편의용. 운영 호스트는 자기 마이그레이션 러너로 같은 파일을 적용한다
        # (db/migrations/README.md).
        applied = 0
        for path in sorted((REPO_ROOT / "db" / "migrations").glob("*.sql")):
            count = apply_sql_file(path, DB_NAME)
            applied += count
            print(f"applied {path.name}: {count} statement(s)")
        print(f"migrations done: {applied} statement(s) total")
        return 0

    book = parse_workbook(EXCEL_PATH)
    print(f"parsed: rows={len(book.rows)} doc_types={len(book.doc_types)}")

    problems = check_consistency(book)
    if problems:
        # 한글 메시지라 콘솔 대신 파일로 남긴다.
        out = REPO_ROOT / "db" / "_migrate_problems.txt"
        out.write_text("\n".join(problems), encoding="utf-8")
        print(f"FAILED: {len(problems)} consistency problem(s) -> {out}")
        return 1
    print("consistency: OK")

    if args.dump:
        io.open(args.dump, "w", encoding="utf-8").write(
            json.dumps(
                {
                    "rows": [vars(r) for r in book.rows],
                    "doc_types": [vars(d) for d in book.doc_types],
                },
                ensure_ascii=False, indent=1, default=str,
            )
        )
        print(f"dumped -> {args.dump}")

    if args.emit_sql:
        emit_seed_sql(book, SEED_OUT_PATH)
        print(f"emitted -> {SEED_OUT_PATH}")

    if args.dry_run:
        print("dry-run: DB untouched")
        return 0

    if not args.skip_schema:
        print(f"schema.sql applied: {apply_sql_file(SCHEMA_PATH, None)} statement(s)")

    stats = load(book)
    print("loaded: " + " ".join(f"{k}={v}" for k, v in stats.items()))

    # **템플릿 적재 뒤에** 적용한다. dev_seed 의 데모 프로젝트는 시드 템플릿의
    # 발행본을 복제하므로, 먼저 돌리면 복제할 원본이 없어 **조용히 빈 프로젝트**가
    # 만들어진다 (실제로 그렇게 만들어 봤다 — 오류 없이 0행이 나온다).
    if args.with_dev_seed:
        print(f"dev_seed.sql applied: {apply_sql_file(DEV_SEED_PATH, DB_NAME)} statement(s)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
