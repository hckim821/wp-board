"""보드 → XLSX — plan.md §0.5.7 (CSV 내보내기를 대체한다).

내보내기는 백엔드가 만든다 (CLAUDE.md 내보내기 원칙). 클라이언트가 그리드를 CSV 로
덤프하는 방식이 아니라 데이터에서 다시 쓰는 방식이라, 화면에 보이지 않는 행도 빠지지
않고 서식이 실행 환경에 좌우되지 않는다.

## 양식은 `docs/Work Package.xlsx` 원본 그대로다

시트 구성·머리말 위치·컬럼 순서·표기 형식·서식(폰트 Malgun Gothic 9.5, 남색 헤더,
얇은 테두리, Phase 색 밴딩, 열 너비, 창 고정)은 **원본 파일을 openpyxl 로 열어 뽑은
실측값**이며 추정이 아니다.

## 내보내기가 곧 임포트 양식이다 — 이것이 이 모듈의 계약이다

완전히 배정된 보드를 내보낸 파일은 `db/migrate.py` 의 `parse_workbook` 이 **그대로
다시 읽을 수 있어야 한다.** 그래서 표기 형식이 취향 문제가 아니다.

* `Phase {n}. {이름}` — `PHASE_RE` 가 요구하는 형태
* `{n}.{m} {이름}` — `MILESTONE_RE` 가 요구하는 형태
* 관련 문서는 `{원문자} {문서명}` 을 ` / ` 로 연결 — 파서는 `/` 가 아니라
  **원문자를 토큰 경계**로 삼는다 (문서명 자체에 `/` 가 들어 있기 때문, §1.1)
* Owner 는 `+` 연결, Status 는 원본 표기(`Not Started` 등)

`tests/test_board_xlsx.py` 의 round-trip 테스트가 이 계약을 fail-closed 로 잠근다 —
표기를 하나라도 흩뜨리면 재파싱이 깨진다.

## 미배정(회색) 행

Phase/Milestone 을 **빈 셀**로 내보낸다. 원본 양식에는 없던 상태이므로 재파싱
대상이 아니다 (`parse_workbook` 은 빈 Phase 셀을 형식 오류로 거부한다). 내보내기는
화면에 있는 것을 그대로 보여주는 쪽을 택하고, 다시 읽을 수 있는지는 발행 검증이
보증하는 "완전히 배정된 보드" 에 대해서만 약속한다.

## openpyxl 은 늦게 import 한다

모듈 최상단에서 import 하면 `create_wp_router()` 가 openpyxl 을 하드 의존하게 된다.
함수 안에서 import 하고, 없으면 이 엔드포인트만 501 로 답한다 (PPTX 경로와 같은 규율).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO

from sqlalchemy.orm import Session

from ..core.exceptions import WpError
from ..models.base import ItemStatus
from ..schemas.item import ItemOut
from .dashboard_theme import phase_color, tint


class ExportUnavailableError(WpError):
    """openpyxl 미설치. **501** — 요청이 틀린 게 아니라 서버가 못 하는 것이다."""

    status_code = 501
    code = "EXPORT_UNAVAILABLE"


# =============================================================================
# 원본 양식 상수 — `docs/Work Package.xlsx` 실측값
# =============================================================================
BOARD_SHEET = "Project Board"
DOC_SHEET = "Doc Status"

BOARD_TITLE = "DSEP AI Project Board (간소화 관리 모판)"
BOARD_NOTE = (
    "관리 원칙: DSEP 인프라 담당자는 개발 세부기술이 아닌 준비상태·일정·요청·이슈·"
    "산출물·평가판정을 관리"
)
#: 헤더가 4행, 데이터가 5행부터 — `parse_workbook` 은 첫 칸이 "No" 인 행을 찾는다.
BOARD_HEADER_ROW = 4
BOARD_FIRST_DATA_ROW = BOARD_HEADER_ROW + 1
BOARD_HEADERS = (
    "No", "Phase", "Milestone", "Key Action Item", "Deliverable (Check Point)",
    "관련 문서", "Owner", "Status", "완료일",
)
BOARD_WIDTHS = {"A": 5.0, "B": 30.0, "C": 26.0, "D": 42.0, "E": 38.0,
                "F": 30.0, "G": 17.0, "H": 12.0, "I": 10.0}
BOARD_FREEZE = f"F{BOARD_FIRST_DATA_ROW}"

DOC_TITLE = "통합 문서 5종 작성 현황"
DOC_HEADER_ROW = 3
#: §0.5.10 — 문서는 순서·이름만 갖는다. 단계/Gate/작성 주체/비고는 전역
#: 문서 모델의 잔재였고 함께 사라졌다.
DOC_HEADERS = ("순서", "문서명")
#: 프로젝트 내보내기에만 붙는 세 컬럼 (§0.5.7 / §0.5-4).
DOC_PROJECT_HEADERS = ("사용", "문서 링크", "작성 상태")
DOC_WIDTHS = {"A": 6.0, "B": 40.0, "C": 8.0, "D": 38.0, "E": 12.0}

FONT_NAME = "Malgun Gothic"
FONT_PT = 9.5
TITLE_PT = 14.0
INK = "FF1E2761"
HEADER_FILL = "FF1E2761"
HEADER_FG = "FFFFFFFF"
TITLE_ROW_H, HEADER_ROW_H, DATA_ROW_H = 20.25, 27.0, 27.0
DATE_FORMAT = "yyyy-mm-dd"

#: enum → **원본 표기**. `migrate.STATUS_MAP` 이 소문자로 되읽으므로 왕복이 닫힌다.
EXPORT_STATUS: dict[ItemStatus, str] = {
    ItemStatus.NOT_STARTED: "Not Started",
    ItemStatus.IN_PROGRESS: "In Progress",
    ItemStatus.DONE: "Done",
    ItemStatus.HOLD: "Hold",
    ItemStatus.NA: "N/A",
}

#: 프로젝트 문서 작성 상태 → 사람이 읽는 표기.
DOC_STATUS_LABEL = {"NOT_WRITTEN": "작성 전", "WRITING": "작성중", "DONE": "완료"}


# =============================================================================
# 셀 값 만들기 — 순수 함수 (openpyxl 없이 테스트된다)
# =============================================================================
def phase_cell(item: ItemOut) -> str:
    """`Phase 0. Pre-Infrastructure Setup`. 미배정이면 빈 문자열."""
    if item.phase_id is None or item.phase_no is None:
        return ""
    return f"Phase {item.phase_no}. {item.phase_name or ''}".rstrip()


def milestone_cell(item: ItemOut) -> str:
    """`0.1 DSEP 환경 Gap 및 자원 구성`. 미배정이면 빈 문자열."""
    if item.milestone_id is None or item.phase_no is None or item.milestone_no is None:
        return ""
    return f"{item.phase_no}.{item.milestone_no} {item.milestone_name or ''}".rstrip()


def documents_cell(item: ItemOut) -> str:
    """`1 / 3` — **파생 표시 번호를 ` / ` 로 연결**한다 (plan.md §0.5.10 ③).

    원문자(①②)는 전역 문서 모델과 함께 폐기됐다. 숫자만 싣는 덕분에 문서명에
    `/` 가 들어 있어도 왕복이 깨질 여지가 아예 없다 — 예전에는 원문자를 토큰
    경계로 삼아 그 함정을 피했다.

    번호는 **사용 중인 문서만 세어** 매긴 값이다 (§0.5.10 팝업 정밀화). 꺼진
    문서는 `build_item_views` 단계에서 이미 빠져 있으므로 여기 들어올 수 없다.
    """
    return " / ".join(str(d.no) for d in item.documents)


def owners_cell(item: ItemOut) -> str:
    return "+".join(o.name for o in item.owners)


@dataclass
class BoardExport:
    """내보낼 한 벌. 시트 두 장에 필요한 것이 전부 들어 있다."""

    title: str
    items: list[ItemOut]
    documents: list = field(default_factory=list)
    #: `document_type_id` → `(사용, 링크, 작성 상태)`. 프로젝트 내보내기에만 있다.
    project_documents: dict[int, tuple[bool, str | None, str]] | None = None


# =============================================================================
# 생성
# =============================================================================
def build_board_xlsx(export: BoardExport) -> bytes:
    """원본 양식 그대로의 통합문서를 바이트로 돌려준다.

    **행이 0개여도 정상적으로 만들어진다** — 머리말과 헤더만 있는 시트가 나온다.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - 설치된 환경에서는 도달하지 않는다
        raise ExportUnavailableError(
            "XLSX 내보내기를 사용하려면 서버에 openpyxl 이 설치되어 있어야 합니다.",
            detail={"package": "openpyxl"},
        ) from exc

    thin = Side(style="thin", color="FFAEB9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    base_font = Font(name=FONT_NAME, size=FONT_PT)
    head_font = Font(name=FONT_NAME, size=FONT_PT, bold=True, color=HEADER_FG)
    head_fill = PatternFill("solid", fgColor=HEADER_FILL)
    head_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="center", wrap_text=True)
    centre = Alignment(horizontal="center", vertical="center", wrap_text=True)

    workbook = Workbook()

    # =========================================================================
    # Project Board
    # =========================================================================
    sheet = workbook.active
    sheet.title = BOARD_SHEET

    sheet["A1"] = BOARD_TITLE
    sheet["A1"].font = Font(name=FONT_NAME, size=TITLE_PT, bold=True, color=INK)
    sheet.row_dimensions[1].height = TITLE_ROW_H
    sheet["A2"] = BOARD_NOTE
    sheet["A2"].font = Font(name=FONT_NAME, size=FONT_PT, color="FF5B6472")

    last_row = BOARD_HEADER_ROW + len(export.items)
    if export.items:
        # 원본의 요약 수식. 행 수가 달라지므로 범위를 다시 만든다.
        summary = {
            "C2": "전체",
            "D2": f"=COUNTA($A${BOARD_FIRST_DATA_ROW}:$A${last_row})",
            "E2": "Done",
            "F2": f'=COUNTIF($H${BOARD_FIRST_DATA_ROW}:$H${last_row},"Done")',
            "G2": "진행률",
            "H2": "=IF(D2=0,0,F2/D2)",
        }
        for address, value in summary.items():
            sheet[address] = value
            sheet[address].font = Font(name=FONT_NAME, size=FONT_PT, bold=True, color=INK)
        sheet["H2"].number_format = "0%"

    for index, header in enumerate(BOARD_HEADERS, start=1):
        cell = sheet.cell(row=BOARD_HEADER_ROW, column=index, value=header)
        cell.font, cell.fill, cell.alignment, cell.border = (
            head_font, head_fill, head_align, border
        )
    sheet.row_dimensions[BOARD_HEADER_ROW].height = HEADER_ROW_H

    for offset, item in enumerate(export.items):
        row_no = BOARD_FIRST_DATA_ROW + offset
        values = (
            item.row_no,
            phase_cell(item),
            milestone_cell(item),
            item.title or "",
            item.deliverable or "",
            documents_cell(item),
            owners_cell(item),
            EXPORT_STATUS.get(ItemStatus(item.status), ""),
            item.completion_date,
        )
        # Phase 색 밴딩 — §0.5 팔레트를 흰색 쪽으로 섞은 연한 색. 미배정은 무색.
        band = (
            PatternFill("solid", fgColor="FF" + tint(phase_color(item.phase_no)))
            if item.phase_id is not None
            else None
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_no, column=column, value=value)
            cell.font = base_font
            cell.border = border
            cell.alignment = centre if column in (1, 8, 9) else body_align
            if column in (2, 3) and band is not None:
                cell.fill = band
            if column == 9:
                cell.number_format = DATE_FORMAT
        sheet.row_dimensions[row_no].height = DATA_ROW_H

    for letter, width in BOARD_WIDTHS.items():
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = BOARD_FREEZE

    # =========================================================================
    # Doc Status
    # =========================================================================
    docs_sheet = workbook.create_sheet(DOC_SHEET)
    docs_sheet["A1"] = DOC_TITLE
    docs_sheet["A1"].font = Font(name=FONT_NAME, size=TITLE_PT, bold=True, color=INK)
    docs_sheet.row_dimensions[1].height = TITLE_ROW_H

    headers = list(DOC_HEADERS)
    if export.project_documents is not None:
        headers += list(DOC_PROJECT_HEADERS)
    for index, header in enumerate(headers, start=1):
        cell = docs_sheet.cell(row=DOC_HEADER_ROW, column=index, value=header)
        cell.font, cell.fill, cell.alignment, cell.border = (
            head_font, head_fill, head_align, border
        )
    docs_sheet.row_dimensions[DOC_HEADER_ROW].height = HEADER_ROW_H

    # ⚠️ **사용(`is_used`) 문서만 싣는다** — §0.5.10 팝업 정밀화가 남긴 선택.
    #
    # 표시 번호가 사용 문서만 세어 매겨지므로, 꺼진 문서를 시트에 함께 실으면
    # 시트의 행 위치와 '관련 문서' 셀의 숫자가 어긋난다 (순서 칸을 `—` 로 비워도
    # 마찬가지다 — 파서는 위치로 문서를 찾는다). 둘 중 하나를 고를 수밖에 없고,
    # **재파싱이 깨지지 않는 쪽**을 골랐다. 꺼진 문서는 화면·전체 현황·PPT 어디에도
    # 나오지 않으므로, 내보내기에서만 보이는 것이 오히려 일관성을 해친다.
    #
    # 대가: 꺼진 문서는 내보낸 파일에 남지 않는다. 다시 켜려면 원본 프로젝트에서
    # 켜야 한다. 왕복 보증은 "사용 중인 문서만" 을 대상으로 한다.
    for offset, document in enumerate(
        [d for d in export.documents if getattr(d, "is_used", 1)]
    ):
        row_no = DOC_HEADER_ROW + 1 + offset
        values = [offset + 1, document.name]
        if export.project_documents is not None:
            used, link, doc_status = export.project_documents.get(
                document.id, (True, None, "NOT_WRITTEN")
            )
            values += [
                "O" if used else "",
                link or "",
                DOC_STATUS_LABEL.get(doc_status, doc_status),
            ]
        for column, value in enumerate(values, start=1):
            cell = docs_sheet.cell(row=row_no, column=column, value=value)
            cell.font = base_font
            cell.border = border
            cell.alignment = centre if column in (1, 3, 5) else body_align

    for index, letter in enumerate(
        (get_column_letter(i + 1) for i in range(len(headers))), start=0
    ):
        width = DOC_WIDTHS.get(letter)
        if width:
            docs_sheet.column_dimensions[letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# =============================================================================
# 조회 → 내보내기
# =============================================================================
def _export_for(session: Session, board, title: str) -> BoardExport:
    """문서는 **보드 스코프**에서 읽는다 (§0.5.10) — 주입받을 전역 마스터가 없다."""
    from sqlalchemy import select

    from . import item_service

    spec = board.spec
    items = item_service.load_ordered_items(session, board)
    documents = list(
        session.scalars(
            select(spec.document_cls)
            .where(spec.master_scope(spec.document_cls) == board.master_scope_id)
            .order_by(spec.document_cls.sort_order, spec.document_cls.id)
        )
    )
    return BoardExport(
        title=title,
        items=item_service.build_item_views(session, board, items),
        documents=documents,
    )


def version_board_xlsx(session: Session, version_id: int) -> tuple[str, bytes]:
    """템플릿 버전 내보내기. 파일명은 `{템플릿명} v{번호}`."""
    from . import version_service

    version = version_service.get_version(session, version_id)
    template = version_service.get_template(session, version.template_id)
    board = version_service.board_of(session, version)
    export = _export_for(session, board, f"{template.name} v{version.version_number}")
    return export.title, build_board_xlsx(export)


def project_board_xlsx(session: Session, project_id: int) -> tuple[str, bytes]:
    """프로젝트 내보내기. Doc Status 에 **사용·링크·작성 상태**가 덧붙는다 (§0.5-4)."""
    from . import project_document_service, project_service

    project = project_service.get_project(session, project_id)
    board = project_service.board_of(project)
    export = _export_for(session, board, project.name)
    export.project_documents = {
        row.id: (bool(row.is_used), row.link_url, row.doc_status.value)
        for row in project_document_service.list_documents(session, project_id)
    }
    return export.title, build_board_xlsx(export)
